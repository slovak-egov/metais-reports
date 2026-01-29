#!/usr/bin/env python3
from __future__ import annotations

import os, sys
import json
import pandas as pd
from datetime import datetime

from tqdm import tqdm

from pathlib import Path
from collections import defaultdict
from typing import Dict, List, Any, Union, Set, TypedDict

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from rapidfuzz import fuzz

from metais.common.date import today_date
from metais.common.project_root import find_project_root
from metais.common.fetch_http import get_json_simple
from metais.common.json_utils import load_json_file, extract_result_array
from metais.fetch.fetch_raw import fetch_simple
from metais.common.worksheet_utils import set_cell_color, set_cell_link, set_cell_frame

Pathish = Union[str, Path, os.PathLike[str]]

_DEFAULT_HEADER = {
    "Accept": "application/json",
    "Content-Type": "application/json",
    "User-Agent": "oCDK_fetcher",
}

_ENT_NAME_KEY = "Gen_Profil_nazov"
_ENT_CODE_KEY = "Gen_Profil_kod_metais"

_C_API_NAME_KEY = "codelistNames"

_ZC_ENT_URL_BASE = "https://metais.slovensko.sk/ci/ZC/"
_C_API_URL_BASE  = "https://metais.slovensko.sk/data-objects/codelists/"

_BRICK_RED = (143, 20, 2)
_BRIGHT_RED = (255, 49, 49)
_HAPPY_GREEN = (88, 224, 106)
_GRAY_BLUE = (176, 196, 222)
_WHITE = (255, 255, 255)

def blend_colors(
    col1: tuple[int, int, int],
    col2: tuple[int, int, int],
    slider: float,
    slider_min: float = 0.0,
    slider_max: float = 100.0,
) -> tuple[int, int, int]:
    if slider_max == slider_min:
        return col2

    # clamp
    if slider <= slider_min:
        fac = 0.0
    elif slider >= slider_max:
        fac = 1.0
    else:
        fac = (slider - slider_min) / (slider_max - slider_min)

    r = round((1 - fac) * col1[0] + fac * col2[0])
    g = round((1 - fac) * col1[1] + fac * col2[1])
    b = round((1 - fac) * col1[2] + fac * col2[2])

    # keep in 0..255 just in case
    r = max(0, min(255, r))
    g = max(0, min(255, g))
    b = max(0, min(255, b))
    return (r, g, b)

Header = Dict[str, Any]
Codelist = Dict[str, Any]

class CodelistEntry(TypedDict):
    header: Header
    codelist: Codelist

def fetch_entity(api_url: str, entity: str, headers: Dict = _DEFAULT_HEADER) -> Any:
    payload = {"parameters": {"target": "nodes", "type": entity, "validOnly": "true"}}

    return fetch_simple(api_url, payload=payload, headers=headers)

def fetch_relation(api_url: str, rel: str, headers: Dict = _DEFAULT_HEADER) -> Any:
    payload = {"parameters": {"target": "relations", "type": rel, "validOnly": "true"}}

    return fetch_simple(api_url, payload=payload, headers=headers)

def fetch_metadata(api_url: str, ent_type: str) -> Any:
    url = api_url.replace("{type}", ent_type)
    return get_json_simple(url)
    
def fetch_codelists_headers_API(base_api_url: str) -> Dict[str, Any]:
    return get_json_simple(base_api_url)

def published_codelist_ids_and_codes(codelists_headers: Dict[str, Any]) -> tuple[list[str], list[str]]:
    published = [h for h in codelists_headers.get("codelists", []) if h.get("codelistState") == "PUBLISHED"]
    ids = [str(h["id"]) for h in published]
    codes = [str(h["code"]) for h in published]
    return ids, codes

def fetch_codelist_header_details_by_ids(ids: list[str], header_detail_api_url: str) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for cid in tqdm(ids, desc="Fetching codelist header details", unit="codelists"):
        out.append(get_json_simple(header_detail_api_url.replace("{id}", cid)))
    return out

def fetch_codelists_by_codes(codes: list[str], detail_api_url: str, filter_items: bool = True) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for code in tqdm(codes, desc="Fetching codelists", unit="codelists"):
        codelist = get_json_simple(detail_api_url.replace("{code}", code))
        if filter_items:
            codelist["codelistsItems"] = [
                it for it in codelist.get("codelistsItems", [])
                if it.get("codelistItemState") == "PUBLISHED"
            ]
        out.append(codelist)
    return out

def key_codelists_by_code(
    codelist_header_details: List[Header],
    codelists: List[Codelist],
) -> Dict[str, CodelistEntry]:
    if len(codelist_header_details) != len(codelists):
        raise ValueError("Headers and codelists should have the same length!")

    res: Dict[str, CodelistEntry] = {}
    for h, c in zip(codelist_header_details, codelists):
        res[h["code"]] = {"header": h, "codelist": c}
    return res

def find_latest_entity(entities: List[Dict[str, Any]]) -> Dict[str, Any]:
    most_recent_date = datetime.fromisoformat("2000-01-01T00:00:00.000")
    most_recent_index = -1
    for i in range(len(entities)):
        ent = entities[i]
        date = datetime.fromisoformat(ent["metaAttributes"]["lastModifiedAt"])
        if date > most_recent_date:
            most_recent_date = date
            most_recent_index = i
    return entities[most_recent_index]

def dedupe_entity_by_latest(entity: List[Dict[str, Any]], attr_name: str) -> List[Dict[str, Any]]:
    dups_by_attr: Dict[str, List[Dict[str, Any]]] = defaultdict(list)

    for ent in entity:
        attr_val = ent.get("attributes", {}).get(attr_name)
        if attr_val is not None:
            dups_by_attr[attr_val].append(ent)

    res: List[Dict[str, Any]] = []
    handled: set[Any] = set()

    for ent in entity:
        attr_val = ent.get("attributes", {}).get(attr_name)

        if attr_val is None:
            res.append(ent)
            continue

        group = dups_by_attr.get(attr_val, [])
        if len(group) <= 1:
            res.append(ent)
            continue

        if attr_val in handled:
            continue
        handled.add(attr_val)

        res.append(find_latest_entity(group))

    return res

def attrs_to_dict(raw_data: Any):
    # convert [{"name":..,"value":..}, ...] to {name:value}
    for entity in raw_data:
        attrs_raw = entity.get("attributes", [])
        if isinstance(attrs_raw, dict):
            continue
        attrs: Dict[str, str] = {}
        for attr in attrs_raw:
            key = attr["name"]
            value = attr["value"]
            attrs[key] = value
        entity["attributes"] = attrs

def find_duplicit_codelists(headers: List[Dict[str, Any]], codelists: List[Dict[str, Any]], ZC: List[Dict[str, Any]]) -> Dict[str, List[str, Any]]:
    duplicities: Dict[str, List[str, Any]] = { "api": defaultdict(list), "entity": defaultdict(list) }

    key_api: Dict[str, List[Any]] = defaultdict(list)

    for i in range(len(headers)):
        h = headers[i]
        c = codelists[i]

        code = h["code"]

        key_api[code].append((h, c))

    key_zc: Dict[str, List[Any]] = defaultdict(list)
        
    for zc in ZC:
        code = zc["attributes"][_ENT_CODE_KEY]

        key_zc[code].append(zc)
    
    for key, cs in key_api.items():
        if len(cs) > 1:
            duplicities["api"][key] = cs
    
    for key, ZC in key_zc.items():
        if len(ZC) > 1:
            duplicities["entity"][key] = ZC

    print(len(duplicities["api"]), len(duplicities["entity"]))

    return duplicities

def get_name(header: Dict[str, Any], key: str = "", lang: str = "sk") -> str:
    if not key:
        return ""
    for name in header[key]:
        if name["language"] == lang:
            return name["value"]
    return ""

def write_table_duplicities(duplicities: Dict[str, List[str, Any]], path: Pathish):
    path = Path(path)

    headers = [
        "Duplicity v api-codelist", "", "",
        "Duplicity v entite ZC", "", "",
    ]

    api_col_code: List[str] = []
    api_col_name: List[str] = []
    api_col_uri:  List[str] = []

    api_col_code.append("Kód číselníka")
    api_col_name.append("Názov")
    api_col_uri.append("URI")

    for code, dups_api in duplicities["api"].items():
        first_row = True
        for rec in dups_api:
            codelist_header_details = rec[0]
            codelist_items = rec[1]

            codelist_name = get_name(codelist_header_details, key=_C_API_NAME_KEY)
            codelist_uri = codelist_header_details["uri"]
            
            api_col_code.append(code if first_row else "")
            api_col_name.append(codelist_name)
            api_col_uri.append(codelist_uri)

            first_row = False

    zc_col_code: List[str] = []
    zc_col_name: List[str] = []
    zc_col_uri:  List[str] = []

    zc_col_code.append("Kód číselníka")
    zc_col_name.append("Názov")
    zc_col_uri.append("URI")

    for code, dups_zc in duplicities["entity"].items():
        first_row = True
        for rec in dups_zc:
            codelist_name = rec["attributes"].get(_ENT_NAME_KEY, "")
            codelist_uri = rec["attributes"].get("Gen_Profil_ref_id", "")
            
            zc_col_code.append(code if first_row else "")
            zc_col_name.append(codelist_name)
            zc_col_uri.append(codelist_uri)

            first_row = False

    max_len = max(
        len(api_col_code),
        len(api_col_name),
        len(api_col_uri),
        len(zc_col_code),
        len(zc_col_name),
        len(zc_col_uri)
    )

    rows: List[List[str]] = []

    for i in range(max_len):
        rows.append(
            [
                api_col_code[i] if i < len(api_col_code) else "",
                api_col_name[i] if i < len(api_col_name) else "",
                api_col_uri[i]  if i < len(api_col_uri)  else "",
                zc_col_code[i]  if i < len(zc_col_code)  else "",
                zc_col_name[i]  if i < len(zc_col_name)  else "",
                zc_col_uri[i]   if i < len(zc_col_uri)   else "",
            ]
        )

    df = pd.DataFrame(rows, columns=headers)

    df.to_excel(path, index=False)

def find_duplicit_codes(codelists_by_code: Dict[str, CodelistEntry]) -> Dict[str, Dict[str, List[Dict[str, Any]]]]:
    duplicities: Dict[str, Dict[str, list[Dict[str, Any]]]] = {}

    for codelist_code, entry in codelists_by_code.items():
        codelist = entry["codelist"]

        codes: DefaultDict[str, List[Dict[str, Any]]] = defaultdict(list)

        for item in codelist.get("codelistsItems", []):
            item_code = item.get("itemCode")
            if item_code is None:
                continue
            codes[str(item_code)].append(item)

        dup_items = {item_code: items for item_code, items in codes.items() if len(items) > 1}
        if dup_items:
            duplicities[codelist_code] = dup_items

    return duplicities

def save_duplicit_code_table(
    duplicities: dict[str, dict[str, list[dict[str, Any]]]],
    codelists_by_code: dict[str, CodelistEntry],
    path: Pathish,
) -> None:
    path = Path(path)

    headers = [
        "Kód číselníka",
        "Názov číselníka",
        "Kód duplicitnej položky číselníka",
        "Názov položky",
    ]

    rows: list[list[str]] = []

    for codelist_code, dup_item_codes in duplicities.items():
        entry = codelists_by_code.get(codelist_code)
        if entry is None:
            codelist_name = ""
        else:
            codelist_name = get_name(entry["header"], key=_C_API_NAME_KEY, lang="sk")

        first_codelist_row = True

        for item_code, dup_items in dup_item_codes.items():
            first_itemcode_row = True

            for item in dup_items:
                item_name = get_name(item, key="codelistItemNames", lang="sk")

                rows.append([
                    codelist_code if first_codelist_row else "",
                    codelist_name if first_codelist_row else "",
                    str(item_code) if first_itemcode_row else "",
                    item_name,
                ])

                first_codelist_row = False
                first_itemcode_row = False

    df = pd.DataFrame(rows, columns=headers)
    df.to_excel(path, index=False)

def connect_entities_to_ZC(entity_by_uuid: Dict[str, Any], je_prepojena_Ciselnik: List[Dict[str, Any]]) -> Dict[str, Any]:
    codelists_by_uuid: Dict[str, List[str]] = {}

    for uuid, entity in entity_by_uuid.items():
        if entity["type"] == "ZC":
            codelists_by_uuid[uuid] = []

    for rel in je_prepojena_Ciselnik:
        ent_uuid = rel["startUuid"]
        codelist_uuid = rel["endUuid"]
        if entity_by_uuid[codelist_uuid]["type"] != "ZC":
            print("Warning! Target entity in je_prepojena_Ciselnik is not of type \"ZC\"")
        if entity_by_uuid[ent_uuid]["type"] == "ZC":
            print("Warning! Source entity in je_prepojena_Ciselnik is of type \"ZC\"")
        codelists_by_uuid[codelist_uuid].append(ent_uuid)
    
    res: Dict[str, Any] = {}

    for codelist_uuid, items_uuid in codelists_by_uuid.items():
        codelist = entity_by_uuid[codelist_uuid]
        code = codelist["attributes"][_ENT_CODE_KEY]
        
        codelist["items"] = []

        for item_uuid in items_uuid:
            entity_rec = entity_by_uuid[item_uuid]

            codelist["items"].append(entity_rec)
        
        res[code] = codelist
            
    return res

def find_matching_codelists(
    codelists_api: Dict[str, Any],
    codelists_entity_full: Dict[str, Any],
) -> Dict[str, Set[str]]:
    buckets: Dict[str, Set[str]] = {
        "api": set(),
        "entity": set(),
        "union": set(),
        "intersection": set(),
    }

    # API codes
    buckets["api"]= codelists_api.keys()

    # Entity codes
    buckets["entity"] = codelists_entity_full.keys()

    buckets["union"] = buckets["api"] | buckets["entity"]
    buckets["intersection"] = buckets["api"] & buckets["entity"]

    # turn "api" and "entity" into "only_api" and "only_entity"
    buckets["api_only"] = buckets["api"] - buckets["intersection"]
    buckets["entity_only"] = buckets["entity"] - buckets["intersection"]

    return buckets

def find_item_overlap(entity: Dict[str, Any], codelist: Dict[str, Any]) -> Tuple[float, float]:
    overlap_by_code = 0.0
    overlap_by_name = 0.0

    if not entity["items"] or len(codelist["codelist"]["codelistsItems"]) == 0:
        return overlap_by_code, overlap_by_name
    
    ent_codes: Set[str] = set()
    ent_names: Set[str] = set()

    for ent in entity["items"]:
        code = ent["attributes"].get(_ENT_CODE_KEY, None)
        name = ent["attributes"].get(_ENT_NAME_KEY, None)

        if code:
            ent_codes.add(code)
        if name:
            ent_names.add(name)

    cl_codes: Set[str] = set()
    cl_names: Set[str] = set()

    for item in codelist["codelist"]["codelistsItems"]:
        code = item["itemCode"]
        name = get_name(item, "codelistItemNames")

        if code:
            cl_codes.add(code)
        if name:
            cl_names.add(name)

    overlap_code_num = len(ent_codes & cl_codes)
    overlap_code_denom = len(ent_codes | cl_codes)

    overlap_name_num = len(ent_names & cl_names)
    overlap_name_denom = len(ent_names | cl_names)

    overlap_by_code = overlap_code_num / overlap_code_denom if overlap_code_denom else 0.0
    overlap_by_name = overlap_name_num / overlap_name_denom if overlap_name_denom else 0.0

    return overlap_by_code, overlap_by_name

def write_table_buckets(
    overlaps: Dict[str, Set[str]],
    codelists_api: Dict[str, Any],
    codelists_entity_full: Dict[str, Any],
    path: Pathish,
):
    path = Path(path)

    headers = [
        "Entitové číselníky",
        "Názov",
        "Poznámka",
        "Číselníky z api/codelist",
        "Názov",
        "Poznámka",
        "Prekryv",
    ]

    N_col = 7

    rows = []

    colors = []

    links = []

    sorted_union = sorted(overlaps["union"])

    for code in sorted_union:
        zc_code = zc_name = zc_note = ""
        api_code = api_name = api_note = ""
        overlap_note = ""

        col_row  = [_BRIGHT_RED]*N_col # by default everything sucks
        link_row = [None]*N_col       # by default no links

        if code in overlaps["entity"]:
            zc_code = code
            zc_name = codelists_entity_full[code]["attributes"].get(_ENT_NAME_KEY, "")
            uuid = codelists_entity_full[code]["uuid"]
            link = _ZC_ENT_URL_BASE + uuid

            col_row[0] = col_row[1] = _WHITE
            link_row[0] = link_row[1] = link

            if not zc_name and len(codelists_entity_full[code]["items"]) == 0:
                zc_note = "Chýba názov a nemá žiadne položky"
            elif len(codelists_entity_full[code]["items"]) == 0:
                zc_note = "Nemá žiadne položky"
            elif not zc_name:
                zc_note = "Chýba názov"
            else:
                zc_note = f"Počet prvkov: {len(codelists_entity_full[code]["items"])}"
                col_row[2] = _WHITE

        if code in overlaps["api"]:
            api_code = code
            api_name = get_name(codelists_api[code]["header"], _C_API_NAME_KEY)
            cid = str(codelists_api[code]["header"]["id"])
            link = _C_API_URL_BASE + cid

            col_row[3] = col_row[4] = _WHITE
            link_row[3] = link_row[4] = link

            if len(codelists_api[code]["codelist"]["codelistsItems"]) == 0 or int(codelists_api[code]["codelist"]["codelistsItemCount"]) == 0:
                api_note = "Nemá žiadne položky"
            else:
                api_note = f"Počet prvkov: {len(codelists_api[code]["codelist"]["codelistsItems"])}"
                col_row[5] = _WHITE
        
        if code in overlaps["api"] and code in overlaps["entity"]:
            score = fuzz.ratio(zc_name, api_name)
            col_name = blend_colors(_GRAY_BLUE, _HAPPY_GREEN, score, slider_min=70, slider_max=100)
            col_row[1] = col_row[4] = col_name

            if len(codelists_entity_full[code]["items"]) > 0 and len(codelists_api[code]["codelist"]["codelistsItems"]) > 0:
                overlap_code, overlap_name = find_item_overlap(codelists_entity_full[code], codelists_api[code])

                col_overlap = blend_colors(_GRAY_BLUE, _HAPPY_GREEN, min(overlap_code, overlap_name), slider_min=0.5, slider_max=1.0)

                col_row[6] = col_overlap

                overlap_note = f"Prekryv podľa kódu: {overlap_code:.3f}, názvu: {overlap_name:.3f}"
            else:
                col_row[6] = _WHITE
        else:
            col_row[6] = _WHITE


        rows.append(
            [
                zc_code, zc_name, zc_note,
                api_code, api_name, api_note,
                overlap_note
            ]
        )

        colors.append(col_row)
        links.append(link_row)

    df = pd.DataFrame(rows, columns=headers)

    df.to_excel(path, index=False)

    wb = load_workbook(path)

    ws = wb.active

    # color missing cells
    for idx in range(len(rows)):
        excel_row = idx + 2
        for col in range(N_col):
            rgb = colors[idx][col]
            url = links[idx][col]

            if url is not None:
                set_cell_link(ws, excel_row, col + 1, url)
            if rgb is not None:
                set_cell_color(ws, excel_row, col + 1, rgb)
            
            set_cell_frame(ws, excel_row, col + 1, "black")
    
    wb.save(path)
    
def write_table_overlap_items(
    codelists_api: Dict[str, Any],  # (actually Dict[str, CodelistEntry], but ok)
    codelists_entity_full: List[Dict[str, Any]],
    code: str,
    path_root: Pathish,
) -> None:
    if code not in codelists_api:
        return

    entity = codelists_entity_full[code]
    
    path_root = Path(path_root)
    path = path_root / f"{code}.xlsx"

    # --- API side ---
    items_api_list = codelists_api[code]["codelist"].get("codelistsItems", [])
    items_api_by_code: Dict[str, Dict[str, Any]] = {}
    for it in items_api_list:
        k = it.get("itemCode")
        if k is None:
            continue
        items_api_by_code[str(k)] = it

    # --- Entity side ---
    items_ent_list = entity.get("items", [])
    items_ent_by_code: Dict[str, Dict[str, Any]] = {}
    for it in items_ent_list:
        k = it.get("attributes", {}).get(_ENT_CODE_KEY)
        if k is None:
            continue
        items_ent_by_code[str(k)] = it

    all_codes: set[str] = set(items_api_by_code.keys()) | set(items_ent_by_code.keys())

    col_ent_code: List[str] = ["Kód položky"]
    col_ent_name: List[str] = ["Názov položky"]
    col_api_code: List[str] = ["Kód položky"]
    col_api_name: List[str] = ["Názov položky"]

    for item_code in sorted(all_codes):
        # entity
        ent_item = items_ent_by_code.get(item_code)
        if ent_item is not None:
            col_ent_code.append(item_code)
            col_ent_name.append(ent_item.get("attributes", {}).get(_ENT_NAME_KEY, ""))
        else:
            col_ent_code.append("")
            col_ent_name.append("")

        # API
        api_item = items_api_by_code.get(item_code)
        if api_item is not None:
            col_api_code.append(item_code)
            col_api_name.append(get_name(api_item, "codelistItemNames", lang="sk"))
        else:
            col_api_code.append("")
            col_api_name.append("")

    rows: List[List[str]] = []
    for i in range(len(col_ent_code)):
        rows.append([col_ent_code[i], col_ent_name[i], col_api_code[i], col_api_name[i]])

    df = pd.DataFrame(
        rows,
        columns=[
            "Položky z entity ZC",
            "",
            "Položky z api-codelists",
            "",
        ],
    )
    df.to_excel(path, index=False)

    wb = load_workbook(path)

    ws = wb.active

    for i, row in enumerate(rows, start = 2):
        if row[0] == "": # no code, didn't find anything
            set_cell_color(ws, i, 1, _BRICK_RED)
            set_cell_color(ws, i, 2, _BRICK_RED)

        if row[2] == "": # no code, didn't find anything
            set_cell_color(ws, i, 3, _BRICK_RED)
            set_cell_color(ws, i, 4, _BRICK_RED)
        
        if row[1] != "" and row[3] != "":
            score = fuzz.ratio(row[1], row[3])
            col = blend_colors(_GRAY_BLUE, _HAPPY_GREEN, score, slider_min=70, slider_max=100)
            set_cell_color(ws, i, 1, col)
            set_cell_color(ws, i, 2, col)
            set_cell_color(ws, i, 3, col)
            set_cell_color(ws, i, 4, col)

    
    wb.save(path)

def write_tables_overlapping(codelists_api: Dict[str, Any], codelists_entity_full: List[Dict[str, Any]], overlaps: Dict[str, Set[str]], path_root: Pathish):
    path_root = Path(path_root)

    codes = overlaps["intersection"]

    for code in codes:
        write_table_overlap_items(codelists_api, codelists_entity_full, code, path_root)

def main():
    report_exec_url = "https://metais.slovensko.sk/api/report/reports/execute/" + os.environ.get("METAIS_REPORT_NUM_PROD", "") + "/type/typ?lang=sk"
    base_api_url = "https://metais.slovensko.sk/api/codelist-repo/codelists/codelistheaders?language=sk&pageNumber=1&perPage=1000"
    detail_api_url = "https://metais.slovensko.sk/api/codelist-repo/codelists/codelistheaders/{code}/codelistitems?language=sk&pageNumber=1&perPage=10000"
    header_detail_api_url = "https://metais.slovensko.sk/api/codelist-repo/codelists/codelistheaders/{id}?lang=sk"
    rel_metadata_api_url = "https://metais.slovensko.sk/api/types-repo/relationshiptypes/relationshiptype/{type}"

    proj_root = find_project_root()
    scratch = proj_root / "scratch"
    dump_dir = scratch / "codelists_assessment"
    date = today_date()
    os.makedirs(dump_dir, exist_ok=True)

    codelists_api_path = dump_dir / f"codelists_api_{date}.json"
    ZC_path = dump_dir / f"ZC_{date}.json" # must be called ZC to actually collide with the broader fetch/load check later
    je_prepojena_Ciselnik_path = dump_dir / f"je_prepojena_Ciselnik_{date}.json"
    je_prepojena_Ciselnik_metadata_path = dump_dir / f"je_prepojena_Ciselnik_metadata_{date}.json"
    table_dup_path = dump_dir / f"codelist_duplicities.xlsx"
    dup_json_path = dump_dir / f"codelist_duplicities.json"
    item_code_dup_path = dump_dir / f"item_codes_duplicities.xlsx"
    item_code_dup_json_path = dump_dir / f"item_codes_duplicities.json"
    codelists_entity_full_path = dump_dir / f"entity_codelist_full_{date}.json"
    api_entity_overlap_path = dump_dir / f"api_entity_overlap.xlsx"
    codelists_items_path_root = dump_dir / "codelist_details"

    summary: Dict[str, Any] = { }

    # fetch ZC entity from the report/execute API, or load from disk if dump exists
    if ZC_path.exists():
        ZC = extract_result_array(load_json_file(ZC_path))
    else:
        ZC = extract_result_array(fetch_entity(report_exec_url, "ZC"))
        attrs_to_dict(ZC)
        ZC_path.write_text(json.dumps(ZC, ensure_ascii=False, indent=2), encoding="utf-8")

    # fetch codelists from the open API, or load from disk if today's dump exists
    if codelists_api_path.exists():
        codelists_api = load_json_file(codelists_api_path)
        duplicities = load_json_file(dup_json_path)
        code_duplicities = load_json_file(item_code_dup_json_path)

        summary["duplicities"] = { "api": duplicities["api"], "entity": duplicities["entity"]}
        summary["itemDuplicities"] = code_duplicities
    else:
        codelists_headers = fetch_codelists_headers_API(base_api_url)
        ids, codes = published_codelist_ids_and_codes(codelists_headers)
        codelist_header_details = fetch_codelist_header_details_by_ids(ids, header_detail_api_url)
        codelists = fetch_codelists_by_codes(codes, detail_api_url)

        codelists_api = key_codelists_by_code(codelist_header_details, codelists)
        
        duplicities = find_duplicit_codelists(codelist_header_details, codelists, ZC)
        dup_json_path.write_text(json.dumps(duplicities, ensure_ascii=False, indent=2), encoding="utf-8")
        write_table_duplicities(duplicities, table_dup_path)

        summary["duplicities"] = { "api": duplicities["api"], "entity": duplicities["entity"]}

        code_duplicities = find_duplicit_codes(codelists_api)
        item_code_dup_json_path.write_text(json.dumps(code_duplicities, ensure_ascii=False, indent=2), encoding="utf-8")
        save_duplicit_code_table(code_duplicities, codelists_api, item_code_dup_path)

        summary["itemDuplicities"] = code_duplicities

        codelists_api_path.write_text(json.dumps(codelists_api, ensure_ascii=False, indent=2), encoding="utf-8")

    ZC = dedupe_entity_by_latest(ZC, _ENT_CODE_KEY)

    # fetch all relations "je_prepojena_Ciselnik" from the report/execute API, or load from disk if dump exists
    if je_prepojena_Ciselnik_path.exists():
        je_prepojena_Ciselnik = extract_result_array(load_json_file(je_prepojena_Ciselnik_path))
    else:
        je_prepojena_Ciselnik = extract_result_array(fetch_relation(report_exec_url, "je_prepojena_Ciselnik"))
        attrs_to_dict(je_prepojena_Ciselnik)
        je_prepojena_Ciselnik_path.write_text(json.dumps(je_prepojena_Ciselnik, ensure_ascii=False, indent=2), encoding="utf-8")

    # fetch "je_prepojena_Ciselnik" metadata from the open API, load from disk if exists
    if je_prepojena_Ciselnik_metadata_path.exists():
        codelists_rels_metadata = load_json_file(je_prepojena_Ciselnik_metadata_path)
    else:
        codelists_rels_metadata = fetch_metadata(rel_metadata_api_url, "je_prepojena_Ciselnik")
        je_prepojena_Ciselnik_metadata_path.write_text(json.dumps(codelists_rels_metadata, ensure_ascii=False, indent=2), encoding="utf-8")

    # find source types for the relation
    sources = [s["technicalName"] for s in codelists_rels_metadata["sources"]]

    # start building entity keyed by uuid:
    entity_by_uuid: Dict[str, Any] = {}
    for rec in ZC:
        uuid = rec["uuid"]
        entity_by_uuid[uuid] = rec

    # fetch all source entities
    for ent in sources:
        ent_path = dump_dir / f"{ent}_{date}.json"
        if ent_path.exists():
            ent_raw_data = extract_result_array(load_json_file(ent_path))
        else:
            ent_raw_data = extract_result_array(fetch_entity(report_exec_url, ent))
            attrs_to_dict(ent_raw_data)
            ent_path.write_text(json.dumps(ent_raw_data, ensure_ascii=False, indent=2), encoding="utf-8")
        
        for rec in ent_raw_data:
            uuid = rec["uuid"]
            entity_by_uuid[uuid] = rec

    if codelists_entity_full_path.exists():
        codelists_entity_full = load_json_file(codelists_entity_full_path)
    else:
        codelists_entity_full = connect_entities_to_ZC(entity_by_uuid, je_prepojena_Ciselnik)
        codelists_entity_full_path.write_text(json.dumps(codelists_entity_full, ensure_ascii=False, indent=2), encoding="utf-8")

    summary["emptyEntityCodelists"] = set()

    for entity in codelists_entity_full.values():
        if len(entity["items"]) == 0:
            summary["emptyEntityCodelists"].add(entity["attributes"][_ENT_CODE_KEY])

    overlaps = find_matching_codelists(codelists_api, codelists_entity_full)

    write_table_buckets(overlaps, codelists_api, codelists_entity_full, api_entity_overlap_path)

    os.makedirs(codelists_items_path_root, exist_ok=True)
    write_tables_overlapping(codelists_api, codelists_entity_full, overlaps, codelists_items_path_root)

    print(f"Duplicity v codelists-api: {len(summary["duplicities"]["api"])}")
    if len(summary["duplicities"]["api"]) > 0:
        for codelist_code, items in summary["duplicities"]["api"].items():
            print(f"   Číselník {codelist_code}:")
            for item in items:
                name = get_name(item["header"], "codelistsNames")
                print(f"      {name}")
    print(f"Duplicity v codelists-entity: {len(summary["duplicities"]["entity"])}")
    if len(summary["duplicities"]["entity"]) > 0:
        for codelist_code, items in summary["duplicities"]["entity"].items():
            print(f"   Číselník {codelist_code}:")
            for item in items:
                name = item["attributes"][_ENT_NAME_KEY]
                uuid = item["uuid"]
                print(f"      {name}, uuid: {uuid}")
    print(f"Duplicity v položkách codelists-api: {len(summary["itemDuplicities"])}")
    if len(summary["itemDuplicities"]) > 0:
        for codelist_code, items in summary["itemDuplicities"].items():
            name = get_name(codelists_api[codelist_code]["header"], "codelistsNames")
            print(f"   Číselník {codelist_code} ({name}):")
            for item_code in items:
                print(f"      {item_code}: {len(items)}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
