#!/usr/bin/env python3
from __future__ import annotations

import sys
import os
import re
import json
import argparse
import requests
import pandas as pd

from os import PathLike
from pathlib import Path
from collections import defaultdict
from typing import Union, List, Dict, Any, Tuple, Iterable, Optional, Set

from rdflib import Graph, URIRef, Literal

from openpyxl import load_workbook
from openpyxl.styles import Font

from metais.common.date import today_date, find_latest_dump
from metais.common.project_root import find_project_root
from metais.common.json_utils import load_json_file, extract_result_array
from metais.fetch.fetch_raw import fetch_simple

Pathish = Union[str, Path, PathLike[str]]
Pair = Tuple[str | None, str | None]

OWL_ONTOLOGY = "http://www.w3.org/2002/07/owl#Ontology"

META_TYPE_MAP = {
    "c_typ_dp.1": "Class",
    "c_typ_dp.2": "ObjectProperty",
    "c_typ_dp.3": "DatatypeProperty",
}


# ---------------------------
# MetaIS fetch & cleanup
# ---------------------------

def fetch_dp(api_url: str) -> Any:
    payload = {"parameters": {"target": "nodes", "type": "DatovyPrvok", "validOnly": "true"}}
    headers = {
        "Accept": "application/json",
        "Content-Type": "application/json",
        "User-Agent": "oCDK_fetcher_DP",
    }
    return fetch_simple(api_url, payload=payload, headers=headers)


def attrs_to_dict(raw_data: Any):
    # convert [{"name":..,"value":..}, ...] to {name:value}
    for entity in raw_data:
        attrs_raw = entity.get("attributes", [])
        if isinstance(attrs_raw, dict):
            continue
        attrs: Dict[str, str] = {}
        for attr in attrs_raw:
            key = attr["name"]
            value = attr["value"]
            attrs[key] = value
        entity["attributes"] = attrs


def validate_attr_injective(raw_data: List[Dict[str, Any]], attr_name: str) -> Dict[str, Any]:
    by_val = defaultdict(list)  # val -> [uuid, uuid, ...]
    problems: Dict[str, Any] = {"missing": [], "duplicate": {}}

    for entity in raw_data:
        uuid = entity.get("uuid")
        val = entity.get("attributes", {}).get(attr_name)

        if val is None:
            problems["missing"].append(uuid)
            continue

        by_val[val].append(uuid)

    for val, uuids in by_val.items():
        if len(uuids) > 1:
            problems["duplicate"][val] = uuids

    if not problems["missing"]:
        problems.pop("missing")
    if not problems["duplicate"]:
        problems.pop("duplicate")

    return problems


def fix_injective_problems(raw_data: List[Dict[str, Any]], attr_problems: Dict[str, Any]) -> List[Dict[str, Any]]:
    if not attr_problems:
        return raw_data

    to_delete: Set[str] = set(attr_problems.get("missing", []))

    dup = attr_problems.get("duplicate", {})
    for _, uuids in dup.items():
        if not uuids:
            continue
        keeper = uuids[-1]
        for u in uuids:
            if u != keeper:
                to_delete.add(u)

    return [e for e in raw_data if e.get("uuid") not in to_delete]


def _is_real_uri(s: str) -> bool:
    return isinstance(s, str) and s.startswith(("http://", "https://"))


def _is_request_element_uri(u: str) -> bool:
    return "uri-registration-reguest-element" in (u or "")


def clean_and_filter_DP(DP_raw: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    Keep only accepted + ensure we have a usable URI in attrs["uri"].
    Rules:
      - accepted: Gen_Profil_RefID_stav_registracie == c_stav_registracie.2
      - choose URI:
          - prefer Gen_Profil_ref_id if it's a real URI and not a request-element
          - else use Profil_DatovyPrvok_kod_datoveho_prvku if it's a real URI
      - if neither gives a URI -> DROP the element
    """
    out: List[Dict[str, Any]] = []

    for elem in DP_raw:
        attrs = elem.get("attributes", {})

        stav = attrs.get("Gen_Profil_RefID_stav_registracie", "")
        if stav != "c_stav_registracie.2":
            continue

        refID = (attrs.get("Gen_Profil_ref_id", "") or "")
        code  = (attrs.get("Profil_DatovyPrvok_kod_datoveho_prvku", "") or "")

        uri = ""
        if _is_real_uri(refID) and not _is_request_element_uri(refID):
            uri = refID
        elif _is_real_uri(code):
            uri = code

        if not uri:
            continue

        attrs["uri"] = uri
        out.append(elem)

    return out


def dict_by_attr(raw_data: List[Dict[str, Any]], attr: str) -> Dict[str, Dict[str, Any]]:
    res: Dict[str, Dict[str, Any]] = {}
    for entry in raw_data:
        attr_val = entry["attributes"][attr]
        res[attr_val] = entry
    return res


# ---------------------------
# CMU RDF parsing
# ---------------------------

def _pred_key(g: Graph, p: URIRef) -> str:
    try:
        q = g.namespace_manager.qname(p)          # e.g. "dct:title", "rdf:type"
        return q.split(":", 1)[1]                # -> "title", "type"
    except Exception:
        return re.split(r"[#/]", str(p))[-1]


def _obj_to_py(o: Any) -> Any:
    if isinstance(o, URIRef):
        return str(o)
    if isinstance(o, Literal):
        try:
            return o.toPython()
        except Exception:
            return str(o)
    return str(o)


def _lang_suffix(o: Any) -> Optional[str]:
    if isinstance(o, Literal):
        lang = (o.language or "").strip()
        if lang:
            return lang.casefold()
    return None


def rdfxml_to_records(path: Pathish) -> List[Dict[str, Any]]:
    """
    Parse RDF/XML into records.

    Enhancement: language-tagged literals become extra keys:
      label_sk, label_en, description_sk, note_sk, ...
    while still preserving aggregate keys (label, description, note, ...).
    """
    g = Graph()
    g.parse(str(path), format="xml")

    subjects = sorted({s for s in g.subjects() if isinstance(s, URIRef)}, key=str)

    out: List[Dict[str, Any]] = []
    for s in subjects:
        attrs_multi: Dict[str, list[Any]] = defaultdict(list)

        for p, o in g.predicate_objects(s):
            key = _pred_key(g, p)
            val = _obj_to_py(o)
            attrs_multi[key].append(val)

            lang = _lang_suffix(o)
            if lang:
                # e.g. label_sk, label_en, description_sk, note_sk, ...
                attrs_multi[f"{key}_{lang}"].append(val)

        attrs: Dict[str, Any] = {k: (v[0] if len(v) == 1 else v) for k, v in attrs_multi.items()}
        attrs["Description"] = str(s)
        out.append({"attributes": attrs})

    return out


def _as_list(x: Any) -> List[Any]:
    if x is None:
        return []
    return x if isinstance(x, list) else [x]


def filter_cmu_records(records: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    for rec in records:
        attrs = rec.get("attributes", {})
        desc = str(attrs.get("Description", ""))

        # 1) only data.gov.sk URIs
        if not (desc.startswith("https://data.gov.sk/") or desc.startswith("http://data.gov.sk/")):
            continue

        # 2) drop ontology headers
        types = {str(t) for t in _as_list(attrs.get("type"))}
        if OWL_ONTOLOGY in types:
            continue

        out.append(rec)
    return out


# ---------------------------
# URI + type normalization / matching
# ---------------------------

def last_token(uri: str) -> str:
    u = (uri or "").rstrip("/")
    if "#" in u:
        return u.rsplit("#", 1)[1]
    if "/" in u:
        return u.rsplit("/", 1)[1]
    return u


def canon_uri_casefold(uri: str) -> str:
    """
    Canonicalize URIs for matching:
      - normalize http://data.gov.sk -> https://data.gov.sk
      - treat /def/ontology/id/... as equivalent to /def/ontology/...
      - strip trailing slash
      - casefold (case-insensitive)
    """
    if not uri:
        return ""
    u = uri.strip()

    # scheme/host normalization
    u = u.replace("http://data.gov.sk/", "https://data.gov.sk/")

    # normalize the weird MetaIS variant (NO double slash)
    u = re.sub(
        r"^https://data\.gov\.sk/def/ontology/id/",
        "https://data.gov.sk/def/ontology/",
        u,
        flags=re.IGNORECASE,
    )
    u = re.sub(
        r"^https://data\.gov\.sk/def/ontology/id#",
        "https://data.gov.sk/def/ontology#",
        u,
        flags=re.IGNORECASE,
    )

    # safety: if anything produced a double slash in this specific area, fix it
    u = u.replace("https://data.gov.sk/def/ontology//", "https://data.gov.sk/def/ontology/")

    u = u.rstrip("/")
    return u.casefold()


def cmu_kind(rec: Dict[str, Any]) -> Optional[str]:
    """
    Return: Class | ObjectProperty | DatatypeProperty | None
    """
    t = rec.get("attributes", {}).get("type")
    if isinstance(t, str):
        t = t.rsplit("#", 1)[-1]
        if t in ("Class", "ObjectProperty", "DatatypeProperty"):
            return t
        return None

    if isinstance(t, Iterable):
        toks = [str(a).rsplit("#", 1)[-1] for a in t]
        if "Class" in toks:
            return "Class"
        if "DatatypeProperty" in toks:
            return "DatatypeProperty"
        if "ObjectProperty" in toks:
            return "ObjectProperty"
    return None


def meta_kind(entity: Dict[str, Any]) -> Optional[str]:
    attrs = entity.get("attributes", {})
    code = attrs.get("Profil_DatovyPrvok_typ_datoveho_prvku", "")
    return META_TYPE_MAP.get(code)


def sort_key_pair(p: Pair) -> Tuple[str, str]:
    u = p[0] or p[1] or ""
    tok = last_token(u)
    return (tok.casefold(), tok)


def build_buckets(
    dp_cmu_by_uri: Dict[str, Dict[str, Any]],
    dp_meta_by_uri: Dict[str, Dict[str, Any]],
) -> Dict[str, Any]:
    """
    Matching key = (canon_uri_casefold(uri), kind)
    So: case-insensitive URI match, but ONLY if types match.
    """
    cmu_norm: dict[Tuple[str, Optional[str]], list[str]] = defaultdict(list)
    meta_norm: dict[Tuple[str, Optional[str]], list[str]] = defaultdict(list)

    for u, rec in dp_cmu_by_uri.items():
        cmu_norm[(canon_uri_casefold(u), cmu_kind(rec))].append(u)

    for u, ent in dp_meta_by_uri.items():
        meta_norm[(canon_uri_casefold(u), meta_kind(ent))].append(u)

    cmu_keys = set(cmu_norm.keys())
    meta_keys = set(meta_norm.keys())
    all_keys  = cmu_keys | meta_keys

    union: List[Pair] = []
    ambiguous: List[Dict[str, Any]] = []

    for k in all_keys:
        a = cmu_norm.get(k, [])
        b = meta_norm.get(k, [])
        if len(a) > 1 or len(b) > 1:
            ambiguous.append({"key": k, "cmu": a, "meta": b})

        if a and b:
            for ua in a:
                for ub in b:
                    union.append((ua, ub))
        elif a:
            for ua in a:
                union.append((ua, None))
        else:
            for ub in b:
                union.append((None, ub))

    union.sort(key=sort_key_pair)
    return {"union": union, "ambiguous": ambiguous}


# ---------------------------
# Possible matches (suggestions)
# ---------------------------

def weak_match(token1: str, token2: str) -> bool:
    a = (token1 or "").casefold()
    b = (token2 or "").casefold()
    return (a in b) or (b in a)


def build_possible_matches(
    uri_cmu: str,
    cmu_type: Optional[str],
    dp_meta_by_uri: Dict[str, Dict[str, Any]],
) -> List[Tuple[str, str]]:
    """
    Return list of (display_name, uuid) suggestions. Unlimited.
    Filter suggestions by same type (if cmu_type known).
    """
    cmu_tok = last_token(uri_cmu)

    out: List[Tuple[str, str]] = []
    for uri_meta, entity in dp_meta_by_uri.items():
        mt = meta_kind(entity)
        if cmu_type and mt and (mt != cmu_type):
            continue

        attrs = entity.get("attributes", {})
        disp = attrs.get("Gen_Profil_anglicky_nazov") or last_token(uri_meta)
        if weak_match(cmu_tok, str(disp)):
            out.append((str(disp), entity.get("uuid", "")))

    out.sort(key=lambda x: (x[0].casefold(), x[0]))
    out = [(n, u) for (n, u) in out if u]
    return out


# ---------------------------
# CMU helpers for Názov / Popis
# ---------------------------

def _first_text(x: Any) -> str:
    if x is None:
        return ""
    if isinstance(x, list):
        for v in x:
            if isinstance(v, str) and v.strip():
                return v.strip()
        return ""
    if isinstance(x, str):
        return x.strip()
    return str(x).strip()


def cmu_nazov(attrs: Dict[str, Any]) -> str:
    # Prefer Slovak label if present; otherwise fall back to whatever label is there.
    return _first_text(attrs.get("label_sk")) or _first_text(attrs.get("label"))


def cmu_popis(attrs: Dict[str, Any]) -> str:
    # Prefer Slovak description/note; then fall back to generic.
    for k in ("description_sk", "note_sk", "comment_sk"):
        v = _first_text(attrs.get(k))
        if v:
            return v
    for k in ("description", "note", "comment"):
        v = _first_text(attrs.get(k))
        if v:
            return v
    return ""


# ---------------------------
# Excel hyperlink helper
# ---------------------------

def set_hyperlink(cell, url: str, link_font: Font):
    cell.hyperlink = url
    cell.font = link_font


# ---------------------------
# main
# ---------------------------

def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(description="Compare CMU RDF list vs MetaIS DPs")

    g = ap.add_mutually_exclusive_group(required=True)
    g.add_argument("--cmu-file", help="Filename under <project_root>/scratch/ with CMU RDF/XML (e.g. egov.rdf).")
    g.add_argument("--cmu-path", help="Full/relative path to CMU RDF/XML.")

    ap.add_argument("--prod", action="store_true", help="Switch to prod")
    ap.add_argument("--test", action="store_true", help="Switch to test")

    args = ap.parse_args(argv)

    if args.prod and args.test:
        raise ValueError("Both --prod and --test were specified; pick one.")

    # Default remains "test" unless --prod is given (same behavior as before)
    if args.prod:
        code = os.environ.get("METAIS_REPORT_NUM_PROD", "")
        if not code:
            raise ValueError("Set METAIS_REPORT_NUM_PROD")
        exec_uri = f"https://metais.slovensko.sk/api/report/reports/execute/{code}/type/typ?lang=sk"
        base_ci = "https://metais.slovensko.sk"
    else:
        code = os.environ.get("METAIS_REPORT_NUM_TEST", "")
        if not code:
            raise ValueError("Set METAIS_REPORT_NUM_TEST")
        exec_uri = f"https://metais-test.slovensko.sk/api/report/reports/execute/{code}/type/typ?lang=sk"
        base_ci = "https://metais-test.slovensko.sk"

    ci_prefix = base_ci.rstrip("/") + "/ci/DatovyPrvok/"

    proj_root = find_project_root()
    scratch = proj_root / "scratch"
    output_root = proj_root / "output"

    # CMU file path
    if args.cmu_path:
        cmu_path = Path(args.cmu_path)
        if not cmu_path.is_absolute():
            cmu_path = (Path.cwd() / cmu_path).resolve()
    else:
        cmu_path = (proj_root / "scratch" / args.cmu_file).resolve()

    if not cmu_path.is_file():
        ap.error(f"CMU file not found: {cmu_path}")

    ontology = cmu_path.stem
    date = today_date()
    meta_DP_path = scratch / f"DP_raw_{date}.json"

    # Load MetaIS DP_raw (cached)
    if meta_DP_path.exists():
        DP_raw = extract_result_array(load_json_file(meta_DP_path))
    else:
        DP_raw = extract_result_array(fetch_dp(exec_uri))
        attrs_to_dict(DP_raw)
        meta_DP_path.write_text(json.dumps(DP_raw, ensure_ascii=False, indent=2), encoding="utf-8")

    # Keep only accepted + give each element attrs["uri"], drop invalids
    DP_raw = clean_and_filter_DP(DP_raw)

    # Ensure uri injective (and dedupe)
    attr_problems = validate_attr_injective(DP_raw, "uri")
    if attr_problems:
        DP_raw = fix_injective_problems(DP_raw, attr_problems)

    # Build lookup by cleaned uri
    dp_meta_by_uri = dict_by_attr(DP_raw, "uri")

    # Load CMU
    CMU_raw = rdfxml_to_records(cmu_path)
    CMU_raw = filter_cmu_records(CMU_raw)
    dp_cmu_by_uri = dict_by_attr(CMU_raw, "Description")

    # Buckets
    buckets = build_buckets(dp_cmu_by_uri, dp_meta_by_uri)

    # Output columns (three main “lists”)
    col_CMU_and_Meta: List[str] = []
    col_CMU_not_Meta: List[str] = []
    col_not_CMU_Meta: List[str] = []
    col_CMU_and_Meta_uuid: List[str] = []

    # NEW: CMU-only details
    col_CMU_not_Meta_nazov: List[str] = []
    col_CMU_not_Meta_popis: List[str] = []

    # For each CMU-not-meta row, store a list of (name, uuid)
    cmu_not_meta_matches: List[List[Tuple[str, str]]] = []

    for uri_cmu, uri_meta in buckets["union"]:
        if uri_cmu and uri_meta:
            col_CMU_and_Meta.append(uri_cmu)
            col_CMU_and_Meta_uuid.append(dp_meta_by_uri[uri_meta]["uuid"])
            continue

        if uri_cmu and not uri_meta:
            rec = dp_cmu_by_uri.get(uri_cmu, {})
            t = cmu_kind(rec)
            if t not in ("Class", "ObjectProperty", "DatatypeProperty"):
                continue

            col_CMU_not_Meta.append(uri_cmu)

            attrs = rec.get("attributes", {})
            col_CMU_not_Meta_nazov.append(cmu_nazov(attrs))
            col_CMU_not_Meta_popis.append(cmu_popis(attrs))

            cmu_not_meta_matches.append(build_possible_matches(uri_cmu, t, dp_meta_by_uri))
            continue

        if (not uri_cmu) and uri_meta:
            # keep only meta URIs belonging to this ontology file
            if f"/{ontology}/" in uri_meta:
                col_not_CMU_Meta.append(uri_meta)
            continue

    # Determine how many "possible match" columns we need (unlimited to the right)
    max_matches_found = max((len(m) for m in cmu_not_meta_matches), default=0)

    # Excel hard limit safety
    MAX_EXCEL_COLS = 16384
    base_cols = 5  # 3 main columns + Názov + Popis
    max_match_cols = min(max_matches_found, MAX_EXCEL_COLS - base_cols)
    if max_matches_found > max_match_cols:
        print(f"WARNING: {max_matches_found} matches needed, but Excel max columns capped us at {max_match_cols}.")

    # Pad 3 main columns to same length
    max_len = max((len(col_CMU_and_Meta), len(col_CMU_not_Meta), len(col_not_CMU_Meta)), default=0)

    col_CMU_and_Meta += [""] * (max_len - len(col_CMU_and_Meta))
    col_CMU_and_Meta_uuid += [""] * (max_len - len(col_CMU_and_Meta_uuid))  # keep in memory only

    col_CMU_not_Meta += [""] * (max_len - len(col_CMU_not_Meta))
    col_not_CMU_Meta += [""] * (max_len - len(col_not_CMU_Meta))

    # NEW: pad Názov/Popis aligned with CMU-not-meta column
    col_CMU_not_Meta_nazov += [""] * (max_len - len(col_CMU_not_Meta_nazov))
    col_CMU_not_Meta_popis += [""] * (max_len - len(col_CMU_not_Meta_popis))

    # Build match columns (values) + hyperlink targets for match cells
    match_values_cols: List[List[str]] = []
    match_link_targets: Dict[Tuple[int, int], str] = {}  # (row_idx0, match_col_idx0) -> url

    for j in range(max_match_cols):
        col_vals = [""] * max_len
        for i in range(min(len(cmu_not_meta_matches), max_len)):
            matches = cmu_not_meta_matches[i]
            if j < len(matches):
                name, uuid = matches[j]
                col_vals[i] = name
                match_link_targets[(i, j)] = ci_prefix + uuid
        match_values_cols.append(col_vals)

    # Headers: matches start after the new Názov/Popis columns
    headers = [
        "CMÚ aj MetaIS",
        "Je v MetaIS, nie je v CMÚ",
        "Je v CMÚ, nie je v MetaIS",
        "Názov",
        "Popis/poznámka",
    ] + (["Možné zhody"] + [""] * (max_match_cols - 1) if max_match_cols > 0 else [])

    # Rows
    rows: List[List[str]] = []
    for i in range(max_len):
        row = [
            col_CMU_and_Meta[i],
            col_not_CMU_Meta[i],
            col_CMU_not_Meta[i],
            col_CMU_not_Meta_nazov[i],
            col_CMU_not_Meta_popis[i],
        ]
        for j in range(max_match_cols):
            row.append(match_values_cols[j][i])
        rows.append(row)

    df = pd.DataFrame(rows, columns=headers)

    out_path = scratch / f"{ontology}.xlsx"
    df.to_excel(out_path, index=False)

    # Postprocess hyperlinks (NO column deletion!)
    wb = load_workbook(out_path)
    ws = wb.active
    link_font = Font(color="0000FF", underline="single")

    def find_col(header_name: str) -> Optional[int]:
        for c in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=c).value == header_name:
                return c
        return None

    col_cmu_and_meta = find_col("CMÚ aj MetaIS")
    col_meta_not_cmu = find_col("Je v MetaIS, nie je v CMÚ")
    col_cmu_not_meta = find_col("Je v CMÚ, nie je v MetaIS")
    col_matches      = find_col("Možné zhody")

    # 1) Column "CMÚ aj MetaIS": hyperlink to MetaIS CI using col_CMU_and_Meta_uuid (memory)
    if col_cmu_and_meta:
        for row in range(2, ws.max_row + 1):
            i0 = row - 2
            uuid = col_CMU_and_Meta_uuid[i0] if i0 < len(col_CMU_and_Meta_uuid) else ""
            if isinstance(uuid, str) and uuid.strip():
                cell = ws.cell(row=row, column=col_cmu_and_meta)
                cell.hyperlink = ci_prefix + uuid.strip()
                cell.font = link_font

    # 2) Column "Je v MetaIS, nie je v CMÚ": hyperlink to MetaIS CI (lookup uuid by URI in dp_meta_by_uri)
    if col_meta_not_cmu:
        for row in range(2, ws.max_row + 1):
            uri = ws.cell(row=row, column=col_meta_not_cmu).value
            if isinstance(uri, str) and uri.startswith(("http://", "https://")):
                ent = dp_meta_by_uri.get(uri)
                if ent and ent.get("uuid"):
                    cell = ws.cell(row=row, column=col_meta_not_cmu)
                    cell.hyperlink = ci_prefix + ent["uuid"]
                    cell.font = link_font

    # 3) Column "Je v CMÚ, nie je v MetaIS": hyperlink to the data.gov.sk URI itself
    if col_cmu_not_meta:
        for row in range(2, ws.max_row + 1):
            uri = ws.cell(row=row, column=col_cmu_not_meta).value
            if isinstance(uri, str) and uri.startswith(("http://", "https://")):
                cell = ws.cell(row=row, column=col_cmu_not_meta)
                cell.hyperlink = uri
                cell.font = link_font

    # 4) Possible matches columns: hyperlink each suggested name to MetaIS CI
    if col_matches and max_match_cols > 0:
        # ensure headers to the right are empty
        for c in range(col_matches + 1, col_matches + max_match_cols):
            ws.cell(row=1, column=c).value = ""

        for (i0, j0), url in match_link_targets.items():
            excel_row = 2 + i0
            excel_col = col_matches + j0
            cell = ws.cell(row=excel_row, column=excel_col)
            if cell.value:
                cell.hyperlink = url
                cell.font = link_font

    wb.save(out_path)

    print(f"Wrote: {out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
