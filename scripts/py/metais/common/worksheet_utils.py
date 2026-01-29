from __future__ import annotations

from typing import Any, Optional

from openpyxl.styles import Font, PatternFill
from openpyxl.styles import Border, Side
from openpyxl.styles.colors import Color
from openpyxl.worksheet.worksheet import Worksheet

__all__ = ["set_cell_color", "set_cell_link", "set_cell_frame"]


def set_cell_color(
    ws: Worksheet,
    row: int,
    col: int,
    rgb: tuple[int, int, int],
) -> None:
    r, g, b = rgb
    argb = f"FF{r:02X}{g:02X}{b:02X}"   # opaque + RGB
    fill = PatternFill(fill_type="solid", fgColor=argb)
    ws.cell(row=row, column=col).fill = fill


def set_cell_link(
    ws: Worksheet,
    row: int,
    col: int,
    url: str,
    override_text: str | None = None,
) -> None:
    cell = ws.cell(row=row, column=col)

    if override_text is not None:
        cell.value = override_text
    else:
        if cell.value in (None, ""):
            cell.value = url

    cell.hyperlink = url
    cell.font = Font(color="0000FF", underline="single")


# --------------------------
# Borders / frames
# --------------------------

_RGB = tuple[int, int, int]

_NAMED_COLORS: dict[str, _RGB] = {
    "black": (0, 0, 0),
    "white": (255, 255, 255),
    "red": (255, 0, 0),
    "green": (0, 128, 0),
    "blue": (0, 0, 255),
    "yellow": (255, 255, 0),
    "orange": (255, 165, 0),
    "purple": (128, 0, 128),
    "gray": (128, 128, 128),
    "grey": (128, 128, 128),
    "lightgray": (211, 211, 211),
    "lightgrey": (211, 211, 211),
    "darkgray": (169, 169, 169),
    "darkgrey": (169, 169, 169),

    "brickred": (143, 20, 2),
    "happygreen": (68, 186, 93),
    "grayblue": (176, 196, 222),
}

def _norm_color_key(s: str) -> str:
    return s.strip().lower().replace(" ", "").replace("_", "").replace("-", "")

def _hex_to_rgb(s: str) -> Optional[_RGB]:
    t = s.strip()
    if t.startswith("#"):
        t = t[1:]
    if len(t) == 3 and all(c in "0123456789abcdefABCDEF" for c in t):
        r = int(t[0] * 2, 16)
        g = int(t[1] * 2, 16)
        b = int(t[2] * 2, 16)
        return (r, g, b)
    if len(t) == 6 and all(c in "0123456789abcdefABCDEF" for c in t):
        r = int(t[0:2], 16)
        g = int(t[2:4], 16)
        b = int(t[4:6], 16)
        return (r, g, b)
    return None

def _parse_color(spec: Any) -> Optional[_RGB]:
    """
    Accept:
      - (r,g,b) ints
      - named colors like "red", "brick_red", "light gray"
      - hex: "#RRGGBB", "RRGGBB", "#RGB"
    Return RGB or None if not a color.
    """
    if _is_rgb(spec):
        return spec

    if isinstance(spec, str):
        rgb = _hex_to_rgb(spec)
        if rgb is not None:
            return rgb
        key = _norm_color_key(spec)
        return _NAMED_COLORS.get(key)

    return None

# OpenPyXL's valid-ish border styles (we'll still accept unknown and fall back)
_VALID_STYLES = {
    "dashdot", "dashdotdot", "dashed", "dotted", "double", "hair",
    "medium", "mediumdashdot", "mediumdashdotdot", "mediumdashed",
    "slantdashdot", "thick", "thin",
}

# Common synonyms / sloppy inputs -> openpyxl style names
_STYLE_ALIASES = {
    "dash-dot": "dashDot",
    "dashdot": "dashDot",
    "dash-dot-dot": "dashDotDot",
    "dashdotdot": "dashDotDot",
    "med": "medium",
    "meddashed": "mediumDashed",
    "mediumdashed": "mediumDashed",
    "medium-dashed": "mediumDashed",
    "mediumdashdot": "mediumDashDot",
    "medium-dash-dot": "mediumDashDot",
    "mediumdashdotdot": "mediumDashDotDot",
    "medium-dash-dot-dot": "mediumDashDotDot",
    "slantdashdot": "slantDashDot",
    "slant-dash-dot": "slantDashDot",
    "dot": "dotted",
    "dots": "dotted",
    "": None,
    "none": None,
    "null": None,
    "no": None,
    "off": None,
    "0": None,
}


def _rgb_to_argb(rgb: _RGB) -> str:
    r, g, b = rgb
    # clamp just in case
    r = max(0, min(255, int(r)))
    g = max(0, min(255, int(g)))
    b = max(0, min(255, int(b)))
    return f"FF{r:02X}{g:02X}{b:02X}"


def _is_rgb(x: Any) -> bool:
    return (
        isinstance(x, (tuple, list))
        and len(x) == 3
        and all(isinstance(v, int) for v in x)
    )


def _normalize_style(style: Any, default: str = "thin") -> str | None:
    if style is None:
        return None
    if not isinstance(style, str):
        return default

    s = style.strip()
    if not s:
        return None

    low = s.lower().replace(" ", "").replace("_", "")
    mapped = _STYLE_ALIASES.get(low, None)

    if mapped is None:
        # not an alias -> try to keep original but normalized for openpyxl
        # openpyxl uses camelCase for some styles; for simple ones lower is fine
        # We'll accept common exact ones:
        if low in _VALID_STYLES:
            # return canonical-ish for openpyxl:
            if low == "dashdot":
                return "dashDot"
            if low == "dashdotdot":
                return "dashDotDot"
            if low == "mediumdashdot":
                return "mediumDashDot"
            if low == "mediumdashdotdot":
                return "mediumDashDotDot"
            if low == "mediumdashed":
                return "mediumDashed"
            if low == "slantdashdot":
                return "slantDashDot"
            return low  # thin/thick/dashed/dotted/double/hair/medium
        # unknown nonsense -> fall back
        return default

    return mapped  # could be None (meaning "no border")


def _make_side(spec: Any, *, default_style: str = "thin") -> Side:
    style: str | None = default_style
    rgb: _RGB | None = None

    # direct color (rgb tuple, hex string, or named string)
    rgb0 = _parse_color(spec)
    if rgb0 is not None:
        rgb = rgb0
        style = default_style

    # direct style (only if it wasn't a color name)
    elif isinstance(spec, str):
        style = _normalize_style(spec, default=default_style)
        rgb = None

    # tuple/list combos
    elif isinstance(spec, (tuple, list)):
        parts = list(spec)
        if len(parts) == 0:
            style = default_style
        elif len(parts) == 1:
            return _make_side(parts[0], default_style=default_style)
        else:
            a, b = parts[0], parts[1]
            a_col = _parse_color(a)
            b_col = _parse_color(b)

            if a_col is not None and isinstance(b, str):
                rgb = a_col
                style = _normalize_style(b, default=default_style)
            elif b_col is not None and isinstance(a, str):
                rgb = b_col
                style = _normalize_style(a, default=default_style)
            else:
                # best-effort: first color anywhere, first style string anywhere
                for p in parts:
                    if rgb is None:
                        pc = _parse_color(p)
                        if pc is not None:
                            rgb = pc
                    if isinstance(p, str) and style == default_style:
                        style = _normalize_style(p, default=default_style)

    else:
        style = default_style
        rgb = None

    if style is None:
        return Side(style=None)

    if rgb is None:
        return Side(style=style)

    return Side(style=style, color=Color(rgb=_rgb_to_argb(rgb)))

def set_cell_frame(
    ws: Worksheet,
    row: int,
    col: int,
    all_sides: Any = None,
    all_sides2: Any = None,
    *,
    left: Any = None,
    right: Any = None,
    top: Any = None,
    bottom: Any = None,
) -> None:
    """
    Set border sides on a cell.

    Positional defaults (apply to ALL sides):
      - set_cell_frame(ws, r, c, "black", "thin")  -> all sides thin black
      - set_cell_frame(ws, r, c, "black")          -> all sides black, default style
      - set_cell_frame(ws, r, c, "thin")           -> all sides thin, default color

    Keyword overrides (apply ONLY to that side, override defaults if provided):
      - set_cell_frame(ws, r, c, right="red")
      - set_cell_frame(ws, r, c, "black", "thin", left=("thick", "red"))
    """
    cell = ws.cell(row=row, column=col)
    border = cell.border or Border()

    updates: dict[str, Side] = {}

    # ---- Apply defaults to all sides if provided ----
    if all_sides is not None:
        base_spec = (all_sides, all_sides2) if all_sides2 is not None else all_sides
        base_side = _make_side(base_spec)  # uses your forgiving parser (named colors too)
        updates.update({
            "left": base_side,
            "right": base_side,
            "top": base_side,
            "bottom": base_side,
        })

    # ---- Per-side overrides ----
    if left is not None:
        updates["left"] = _make_side(left)
    if right is not None:
        updates["right"] = _make_side(right)
    if top is not None:
        updates["top"] = _make_side(top)
    if bottom is not None:
        updates["bottom"] = _make_side(bottom)

    if not updates:
        return

    try:
        cell.border = border.copy(**updates)
    except Exception:
        cell.border = Border(
            left=updates.get("left", border.left),
            right=updates.get("right", border.right),
            top=updates.get("top", border.top),
            bottom=updates.get("bottom", border.bottom),
            diagonal=getattr(border, "diagonal", None),
            diagonalUp=getattr(border, "diagonalUp", False),
            diagonalDown=getattr(border, "diagonalDown", False),
            outline=getattr(border, "outline", True),
            vertical=getattr(border, "vertical", None),
            horizontal=getattr(border, "horizontal", None),
        )
