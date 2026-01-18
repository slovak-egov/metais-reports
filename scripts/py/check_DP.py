#!/usr/bin/env python3
import re
import json
import pandas as pd
from pathlib import Path
from os import PathLike
from typing import Any, Mapping, Union, List, Dict, Tuple
from urllib.parse import urlparse
from datetime import date, datetime
from openpyxl import load_workbook
from rapidfuzz import process, fuzz
from collections import defaultdict
from copy import copy
from openpyxl.styles import Color, PatternFill, Border, Side
from openpyxl.cell.rich_text import CellRichText, TextBlock, InlineFont

from metais.common.date import find_latest_dump
from metais.packed_reader.packed_reader import PackedReader
from metais.common.json_utils import load_json_file
from metais.common.project_root import find_project_root

Pathish = Union[str, Path, PathLike[str]]


# ----------------- small utils -----------------

def get_lines(filename: Pathish) -> List[str]:
    with open(filename, "r", encoding="utf-8") as f:
        return [line.rstrip("\n") for line in f]

def name_from_url(s: Any) -> str:
    if not isinstance(s, str):
        return ""
    s = s.strip()
    if not s:
        return ""
    p = urlparse(s)
    path = p.path if (p.scheme or p.netloc) else s  # plain token/path
    return path.rstrip("/").split("/")[-1]

def cellify(v: Any) -> Any:
    """Excel-friendly cell value."""
    if v is None:
        return ""
    if isinstance(v, (date, datetime)):
        return v.isoformat()
    if isinstance(v, list):
        parts = []
        for x in v:
            cx = cellify(x)
            if cx == "" or cx is None:
                continue
            parts.append(str(cx))
        return "; ".join(parts)
    if isinstance(v, dict):
        # shouldn't happen if enum_mode="value", but just in case
        return json.dumps(v, ensure_ascii=False)
    return v


# ----------------- fuzzy helpers -----------------

_STOP: set[str] = set()  # keep empty to avoid collapsing things like VAT->vatID
_RX_CAMEL = re.compile(r"(?<=[a-z0-9])(?=[A-Z])|(?<=[A-Z])(?=[A-Z][a-z])")

def id_tokens(s: str) -> list[str]:
    s = (s or "").strip().rstrip("/")
    s = re.sub(r"[_\-\s]+", " ", s)
    out: list[str] = []
    for chunk in s.split():
        chunk = _RX_CAMEL.sub(" ", chunk)
        out.extend(chunk.split())
    toks = [t.lower() for t in out if t]
    if _STOP:
        toks = [t for t in toks if t not in _STOP]
    return toks

def id_processor(s: str) -> str:
    return " ".join(id_tokens(s))

def _norm_first_tok(t: str) -> str:
    t = t.lower()
    # cheap plural handling: contracts -> contract
    if len(t) > 3 and t.endswith("s"):
        t = t[:-1]
    return t

def accept_match(q: str, m: str, *, first_tok_min_ratio: int = 85) -> bool:
    tq = id_tokens(q)
    tm = id_tokens(m)
    if not tq or not tm:
        return False

    a = _norm_first_tok(tq[0])
    b = _norm_first_tok(tm[0])

    # allow small differences
    return fuzz.ratio(a, b) >= first_tok_min_ratio


# ----------------- report builder -----------------

def build_attr_columns(attr_meta: Any) -> Tuple[List[str], Dict[str, str]]:
    """
    Returns:
      - ordered technicalNames (initial order from meta file)
      - map technicalName -> unique human column label
    """
    if not isinstance(attr_meta, list):
        raise TypeError("attributes.json must be a JSON list")

    techs: list[str] = []
    tech_to_human: dict[str, str] = {}

    used: set[str] = set()

    for x in attr_meta:
        if not isinstance(x, dict):
            continue
        tn = x.get("technicalName")
        hn = x.get("name")
        if not isinstance(tn, str) or not tn:
            continue
        if not isinstance(hn, str) or not hn:
            hn = tn

        col = hn
        if col in used:
            col = f"{hn} ({tn})"
        used.add(col)

        techs.append(tn)
        tech_to_human[tn] = col

    return techs, tech_to_human

def _hex_to_rgb(hex6: str) -> tuple[int, int, int]:
    hex6 = hex6.strip().lstrip("#")
    return (int(hex6[0:2], 16), int(hex6[2:4], 16), int(hex6[4:6], 16))

def _rgb_to_argb_hex(r: int, g: int, b: int) -> str:
    return f"FF{r:02X}{g:02X}{b:02X}"

def _lerp(a: int, b: int, t: float) -> int:
    return int(round(a + (b - a) * t))

def _blend_hex(gray_hex: str, blue_hex: str, t: float) -> str:
    r1, g1, b1 = _hex_to_rgb(gray_hex)
    r2, g2, b2 = _hex_to_rgb(blue_hex)
    r = _lerp(r1, r2, t)
    g = _lerp(g1, g2, t)
    b = _lerp(b1, b2, t)
    return _rgb_to_argb_hex(r, g, b)

def _norm_url_cell(v: Any) -> list[str]:
    """
    Cell may contain:
      - "" (because you blank duplicates)
      - a single URL
      - multiple URLs joined by '; '
    Return normalized list of urls for comparison.
    """
    if v is None:
        return []
    s = str(v).strip()
    if not s:
        return []
    parts = [p.strip() for p in s.split(";") if p.strip()]
    # normalize (ignore trailing slash)
    return [p.rstrip("/") for p in parts]

def _norm_one(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip().rstrip("/")

def _common_prefix_len(a: str, b: str) -> int:
    a2 = a.lower()
    b2 = b.lower()
    n = min(len(a2), len(b2))
    i = 0
    while i < n and a2[i] == b2[i]:
        i += 1
    return i

_BLACK = Color(rgb="FF000000")  # ARGB black

def _force_black(cell) -> None:
    f = copy(cell.font)
    f.color = _BLACK
    cell.font = f

def _set_richtext(cell, segments: list[tuple[str, bool]], *, force_black: bool = True) -> None:
    """
    segments: [(text, bold), ...]
    """
    rt = CellRichText()
    for txt, is_bold in segments:
        if not txt:
            continue
        if force_black:
            font = InlineFont(b=bool(is_bold), color=_BLACK)
        else:
            font = InlineFont(b=bool(is_bold))
        rt.append(TextBlock(font, txt))
    cell.value = rt

    # extra safety: set cell-level font color too (helps some renderers)
    if force_black:
        _force_black(cell)

def format_excel(
    path,
    *,
    col_meta: str,
    col_score: str,
    col_status: str = "Stav registrácie",
    accepted_status: str = "Akceptovaná registrácia",
    col_refid: str = "Referencovateľný identifikátor",
    col_cmu_uri: str = "URI z CMÚ",
    happy_green_hex: str = "11db4a",
    unsure_green_hex: str = "9ae7b2",
    id_match_hex: str = "00B050",
    min_score: int = 70,
    max_score: int = 99,
    blue_hex: str = "a1dde2",
    dark_blue_hex: str = "ccdeda",
    gray_hex: str = "dbdbdb",
) -> None:
    wb = load_workbook(path)
    ws = wb.active

    headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}

    c_meta   = headers.get(col_meta)
    c_score  = headers.get(col_score)
    c_status = headers.get(col_status)
    c_refid  = headers.get(col_refid)
    c_uri    = headers.get(col_cmu_uri)

    if c_meta is None or c_score is None:
        raise KeyError(f"Missing expected columns: {col_meta=} {col_score=}")

    fill_happy = PatternFill("solid", fgColor=_rgb_to_argb_hex(*_hex_to_rgb(happy_green_hex)))
    fill_unsure = PatternFill("solid", fgColor=_rgb_to_argb_hex(*_hex_to_rgb(unsure_green_hex)))
    fill_gray  = PatternFill("solid", fgColor=_rgb_to_argb_hex(*_hex_to_rgb(gray_hex)))
    fill_idmatch = PatternFill("solid", fgColor=_rgb_to_argb_hex(*_hex_to_rgb(id_match_hex)))

    thin = Side(style="thin", color="1e1e1e")
    grid_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    lo = int(min_score)
    hi = int(max_score) if max_score > min_score else (min_score + 1)

    last_uri_cell_value = ""  # for duplicate-blanked rows

    for r in range(1, ws.max_row + 1):
        row_fill = None

        if r >= 2:
            meta_val = ws.cell(row=r, column=c_meta).value
            score_val = ws.cell(row=r, column=c_score).value

            # keep last non-empty URI
            if c_uri is not None:
                uri_cell = ws.cell(row=r, column=c_uri).value
                if uri_cell is not None and str(uri_cell).strip() != "":
                    last_uri_cell_value = str(uri_cell)

            if meta_val is None or str(meta_val).strip() == "":
                row_fill = fill_gray
            else:
                try:
                    s = int(score_val)
                except Exception:
                    s = None

                if s == 100:
                    # choose green by status (if status column exists)
                    status_ok = False
                    if c_status is not None:
                        st = ws.cell(row=r, column=c_status).value
                        status_ok = (str(st).strip() == accepted_status) if st is not None else False
                    row_fill = fill_happy if status_ok else fill_unsure

                elif s is not None:
                    s_clamped = max(lo, min(hi, s))
                    t = (s_clamped - lo) / (hi - lo)
                    color_argb = _blend_hex(dark_blue_hex, blue_hex, t)
                    row_fill = PatternFill("solid", fgColor=color_argb)

        # apply borders + row fill + force font color
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = grid_border
            if row_fill is not None:
                cell.fill = row_fill
                if r >= 2:   # keep header as-is
                    _force_black(cell)

        # highlight + bold common prefix between (Referencovateľný identifikátor) and (URI z CMÚ)
        if r >= 2 and c_refid is not None and c_uri is not None:
            ref_cell = ws.cell(row=r, column=c_refid)
            uri_cell = ws.cell(row=r, column=c_uri)

            ref_raw = ref_cell.value
            refid_val = _norm_one(ref_raw)  # normalized (rstrip("/"), strip)

            # use last known URI if this row’s URI cell is blank (because you blank duplicates)
            uri_val_raw = uri_cell.value
            uri_val_effective = str(uri_val_raw).strip() if (uri_val_raw is not None) else ""
            if not uri_val_effective:
                uri_val_effective = last_uri_cell_value

            uris = _norm_url_cell(uri_val_effective)

            if refid_val and uris:
                # choose the CMU URI that shares the longest prefix with the refid
                best_uri = max(uris, key=lambda u: _common_prefix_len(u, refid_val))
                L = _common_prefix_len(best_uri, refid_val)

                # 1) cell background highlight when it is an exact match (your old rule)
                if refid_val in uris:
                    ref_cell.fill = fill_idmatch
                    uri_cell.fill = fill_idmatch

                # 2) bold the common prefix (only if it's "meaningful")
                # avoids bolding trivial "https://"
                MIN_BOLD_PREFIX = 20
                if L >= MIN_BOLD_PREFIX:
                    # --- RefID cell: bold prefix from start ---
                    ref_text = "" if ref_raw is None else str(ref_raw)
                    if ref_text:
                        L_ref = min(L, len(ref_text))
                        _set_richtext(ref_cell, [
                            (ref_text[:L_ref], True),
                            (ref_text[L_ref:], False),
                        ])

                    # --- URI cell: bold prefix only inside the matching URL substring ---
                    # If the URI cell is blank (duplicate rows), don't try to rich-text it
                    uri_text = "" if uri_val_raw is None else str(uri_val_raw)
                    if uri_text.strip():
                        # find the best_uri inside the displayed cell text (may contain multiple URLs)
                        # (best_uri has rstrip("/") normalization; search both variants)
                        idx = uri_text.find(best_uri)
                        if idx < 0:
                            idx = uri_text.find(best_uri + "/")
                        if idx >= 0:
                            # determine the actual substring in the cell we're bolding against
                            # (use the exact slice from uri_text to keep case/slashes)
                            # take until next ';' or end
                            end = uri_text.find(";", idx)
                            if end < 0:
                                end = len(uri_text)
                            chosen = uri_text[idx:end].strip()
                            L_uri = min(_common_prefix_len(chosen.rstrip("/"), refid_val), len(chosen))

                            pre = uri_text[:idx]
                            post = uri_text[end:]
                            _set_richtext(uri_cell, [
                                (pre, False),
                                (chosen[:L_uri], True),
                                (chosen[L_uri:], False),
                                (post, False),
                            ])

    wb.save(path)


def main():
    proj_root = find_project_root()
    date_str = find_latest_dump(proj_root / "output")

    urls = get_lines(proj_root / "scratch" / "fin_CMU_new.txt")

    cmu_urls_by_name: dict[str, list[str]] = defaultdict(list)
    for u in urls:
        n = name_from_url(u)
        if n:
            cmu_urls_by_name[n].append(u)

    attr_meta = load_json_file(
        proj_root / "output" / date_str / "packed" / "nodes" / "DatovyPrvok" / "attributes.json"
    )

    # CMU tokens
    DP_CMU = sorted({name_from_url(u) for u in urls if name_from_url(u)})

    # attribute column definitions
    tech_names, tech_to_human = build_attr_columns(attr_meta)

    # Load MetaIS DatovyPrvok profiles + attrs
    meta_profiles: list[str] = []
    meta_records: list[dict[str, Any]] = []
    index_by_profile: dict[str, list[int]] = {}

    with PackedReader(
        date=date_str,
        dict_cache_size=None,
        attr_cache_size=None,
        resolver_cache_size=None,
        open_relation_partitions_max=None,
    ) as pr:
        for DP in pr.iterate_citype(
            "DatovyPrvok",
            include_attrs=True,
            include_meta=True,
            valid_only=True,
        ):
            # to unpack (attr, meta_attr) we must use meta_prefix=None
            attr, meta_attr = pr.get_attributes_typed(
                DP,
                include_meta=True,
                meta_prefix=None,
                enum_mode="value",   # use human-readable labels from enums
                return_info=False,
            )

            profil = name_from_url(attr.get("Profil_DatovyPrvok_kod_datoveho_prvku", ""))
            if not profil:
                continue

            idx = len(meta_profiles)
            meta_profiles.append(profil)
            meta_records.append(attr)

            index_by_profile.setdefault(profil, []).append(idx)

    # --- Build rows: one per (CMU, MetaIS match) ---
    score_cutoff = 70
    top_k = 25  # max candidates per CMU

    # Pass 1: gather MetaIS tokens that have a 100% match anywhere
    claimed_100_meta: set[str] = set()

    for cmu in DP_CMU:
        # exact 100% claims
        for idx in index_by_profile.get(cmu, []):
            claimed_100_meta.add(meta_profiles[idx])

        # fuzzy 100% claims (still guarded)
        hits = process.extract(
            cmu,
            meta_profiles,
            scorer=fuzz.token_sort_ratio,
            processor=id_processor,
            score_cutoff=100,
            limit=top_k,
        )
        for match, score, _idx in hits:
            if int(score) == 100 and accept_match(cmu, match):
                claimed_100_meta.add(match)

    # Pass 2: actually emit rows (100s first, then <100 excluding claimed_100_meta)
    rows_raw: list[dict[str, Any]] = []

    for cmu in DP_CMU:
        emitted_meta: set[str] = set()  # per-CMU dedupe

        # 1) exact matches (100)
        for idx in index_by_profile.get(cmu, []):
            m = meta_profiles[idx]
            rows_raw.append({"cmu": cmu, "meta": m, "score": 100, "attrs": meta_records[idx]})
            emitted_meta.add(m)

        # 2) fuzzy matches (>= cutoff), one row per match
        hits = process.extract(
            cmu,
            meta_profiles,
            scorer=fuzz.token_sort_ratio,
            processor=id_processor,
            score_cutoff=score_cutoff,
            limit=top_k,
        )

        for match, score, idx in hits:
            score_i = int(score)
            if not accept_match(cmu, match):
                continue
            if match in emitted_meta:
                continue

            # KEY RULE:
            # if this MetaIS token is already matched with 100 elsewhere,
            # don't use it as a weaker (<100) match for this CMU token.
            if score_i < 100 and match in claimed_100_meta:
                continue

            rows_raw.append({"cmu": cmu, "meta": match, "score": score_i, "attrs": meta_records[int(idx)]})
            emitted_meta.add(match)

        # 3) no matches -> still a row
        if not emitted_meta:
            rows_raw.append({"cmu": cmu, "meta": "", "score": "", "attrs": {}})

    # Sort rows: group by CMU, then score desc
    def _score_key(x: Any) -> int:
        try:
            return int(x)
        except Exception:
            return -1

    rows_raw.sort(key=lambda r: (r["cmu"].lower(), -_score_key(r["score"]), str(r["meta"]).lower()))

    # Count attribute frequency across rows (only when a MetaIS record exists)
    freq: dict[str, int] = {tn: 0 for tn in tech_names}
    for r in rows_raw:
        attrs = r.get("attrs") or {}
        if not attrs:
            continue
        for tn in tech_names:
            v = attrs.get(tn)
            cv = cellify(v)
            if cv != "":
                freq[tn] += 1

    tech_sorted = sorted(
        tech_names,
        key=lambda tn: (-freq.get(tn, 0), tech_to_human.get(tn, tn).lower())
    )

    # Build final dataframe rows
    col_url  = "URI z CMÚ"
    col_cmu = "Názov prvku z CMÚ"
    col_meta = "Zhodný prvok z MetaIS?"
    col_score = "Fuzzy match (%)"

    final_rows: list[dict[str, Any]] = []
    for r in rows_raw:
        attrs = r.get("attrs") or {}
        out: dict[str, Any] = {
            col_url: "; ".join(cmu_urls_by_name.get(r["cmu"], [])),
            col_cmu: r["cmu"],
            col_meta: r["meta"],
            col_score: r["score"],
        }
        for tn in tech_sorted:
            col = tech_to_human.get(tn, tn)
            out[col] = cellify(attrs.get(tn))
        final_rows.append(out)

    df = pd.DataFrame(final_rows)

    df[col_url] = df[col_url].mask(df[col_cmu].duplicated(), "")
    df[col_cmu] = df[col_cmu].mask(df[col_cmu].duplicated(), "")

    # Save
    out_path = proj_root / "scratch" / "dp_cmu_vs_metais.xlsx"
    df.to_excel(out_path, index=False)

    format_excel(
        out_path,
        col_meta="Zhodný prvok z MetaIS?",
        col_score="Fuzzy match (%)",
        col_status="Stav registrácie",
        accepted_status="Akceptovaná registrácia",
        col_refid="Referencovateľný identifikátor",
        col_cmu_uri="URI z CMÚ",
        min_score=70,
    )

    print(f"Wrote: {out_path}")
    print(f"Rows: {len(df)}   Columns: {len(df.columns)}")


if __name__ == "__main__":
    main()